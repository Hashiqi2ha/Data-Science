# -*- coding:utf-8 -*-
from pyxlsb import open_workbook as xlsb
import pymysql
import re
import datetime
from openpyxl import load_workbook


start = datetime.datetime.now()
print('开始时间：' + str(start))

db_config = {
    'host': '127.0.0.1',
    'port': 3306,
    'user': 'root',
    'password': '1234',
    'db': 'work',
    'charset': 'utf8',

}
conn = pymysql.connect(**db_config)
cursor = conn.cursor()


def insert_many_sql(_db_name='', key_list=[]):
    fields_sql = '`,`'.join(key_list)
    values_sql = ','.join((u'%s'.format(t) for t in key_list))
    insert_sql = u'insert ignore into `{}` (`{}`) VALUES ({})'.format(_db_name, fields_sql, values_sql)
    return insert_sql


wb = load_workbook(r'C:\Users\65748\OneDrive\桌面\work\EI.xlsx')
sheets = wb.sheetnames
serials = wb['SERIALS']
all_list = []
for row in serials:
    li = []
    for cell in row:
        li.append(cell.value)
    all_list.append(li)

key_list = eval(
    str(all_list[1]).replace(r"Open Access (OA, verified as of Mar. 1, 2021, see definition)\n(开源期刊，2021年3月1日确认)",
                             "Open Access"))
data_list = all_list[2:]
pos_ISSN = key_list.index('ISSN')
pos_EISSN = key_list.index('EISSN')
new_data_list = []
for xs in data_list:
    li_ = []
    i = -1
    for x in xs:
        i += 1
        if i == pos_ISSN or i == pos_EISSN:
            if x == '-':
                x = ''
            else:
                x = str(x).zfill(8)
        else:
            if x =='-':
                x = ''
            else:
                pass
        li_.append(x)
    new_data_list.append(li_)

# print(new_data_list)
# exit()

sql = insert_many_sql('EI', key_list)
cursor.executemany(sql, new_data_list)
conn.commit()
end = datetime.datetime.now()
print('\n' + '结束时间：' + str(end) + '\n')
print('运行时间：' + str(end - start))
